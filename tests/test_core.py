from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from sales_report.config import load_config
from sales_report.analysis import _detect_recipient_mix_shift, _status_for_monthly, _status_for_ytd
from sales_report.input_loader import (
    CUMULATIVE_REQUIRED_HEADERS,
    RECIPIENT_REQUIRED_HEADERS,
    _find_header,
    _normalize_recipient_name,
    _unit_divisors,
    load_cumulative_datasets,
    load_recipient_datasets,
)
from sales_report.models import InputManifest, OfficialMetric, RawSheetCandidate
from sales_report.narrative import (
    _focus_paragraphs,
    _number_destination_particle,
    _recipient_mix_shift_paragraphs,
    _secondary_recipient_paragraphs,
    _status_particle,
    _trend_details,
)
from sales_report.utils import normalize_period, pct_change, previous_period, safe_unit_price


ROOT = Path(__file__).resolve().parents[1]


class UtilityTests(unittest.TestCase):
    def test_previous_period_crosses_year(self):
        self.assertEqual(previous_period("202601"), "202512")
        self.assertEqual(previous_period("202606"), "202605")

    def test_normalize_period(self):
        self.assertEqual(normalize_period("2026-06-30"), "202606")
        self.assertEqual(normalize_period(202606), "202606")

    def test_safe_ratios(self):
        self.assertAlmostEqual(pct_change(120, 100), 0.2)
        self.assertIsNone(pct_change(120, 0))
        self.assertEqual(safe_unit_price(300, 100), 3)
        self.assertIsNone(safe_unit_price(300, 0))

    def test_korean_particles(self):
        self.assertEqual(_number_destination_particle(2.23), "으로")
        self.assertEqual(_number_destination_particle(2.19), "로")
        self.assertEqual(_number_destination_particle(1.88), "로")
        self.assertEqual(_status_particle("출고 재개"), "로")
        self.assertEqual(_status_particle("당월 미출고"), "로")


class ConfigTests(unittest.TestCase):
    def test_config_has_all_report_segments(self):
        config = load_config(ROOT / "config" / "analysis_rules.yaml")
        self.assertEqual(len(config["report_order"]), 6)
        self.assertTrue(set(config["report_order"]).issubset(config["segments"]))
        self.assertIn("synthetic_export", config["report_order"])
        self.assertEqual(config["excluded_accounts"], [6])
        self.assertEqual(config["segments"]["synthetic_domestic"]["primary_level"], 1)
        self.assertEqual(config["segments"]["stainless_export"]["primary_level"], 1)
        self.assertEqual(config["segments"]["steel_export"]["primary_level"], 2)
        self.assertEqual(config["segments"]["steel_export"]["secondary_level"], 2)
        self.assertEqual(config["segments"]["steel_domestic"]["secondary_level"], 2)
        mix_rules = config["output"]["recipient_mix_shift"]
        self.assertEqual(mix_rules["comparisons"], ["전월 대비"])
        self.assertEqual(mix_rules["minimum_side_amount"]["domestic"], 100)


class RecipientPatternTests(unittest.TestCase):
    def test_new_recipient_uses_prior_year_history(self):
        item = {"previous_amount": 0, "current_amount": 120, "amount_delta": 120}
        self.assertEqual(_status_for_monthly(item, {"202607": 120}, "202607"), "신규 실적")
        self.assertEqual(_status_for_monthly(item, {"202507": 50, "202607": 120}, "202607"), "전월 미출고 후 당월 실적")

    def test_ytd_status(self):
        self.assertEqual(_status_for_ytd({"previous_amount": 0, "current_amount": 10, "amount_delta": 10}), "당해 신규")
        self.assertEqual(_status_for_ytd({"previous_amount": 10, "current_amount": 0, "amount_delta": -10}), "당해 미출고")


class CumulativeRuleTests(unittest.TestCase):
    def _make_source(self, directory: Path, year: int) -> RawSheetCandidate:
        headers = [
            "부문", "내수/수출", "요청월", "레벨1명", "레벨2명", "레벨3명", "계정",
            "중량", "원화금액", "달러금액", "한국원화금액", "담당자명", "담당자(세부)명",
        ]
        path = directory / f"{year}7월누계.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(headers)
        worksheet.append(["제강", "내수", f"{year}07", "스틸로프", "WIRE ROPE", "WRG.I", 2, 1000, 1_000_000, 0, 1_000_000, "내수", "내수"])
        worksheet.append(["제강", "내수", f"{year}07", "스틸상품", "스틸상품", "상품", 2, 2000, 2_000_000, 0, 2_000_000, "내수", "내수"])
        worksheet.append(["제강", "내수", f"{year}07", "", "", "", 6, 3000, 3_000_000, 0, 3_000_000, "내수", "내수"])
        workbook.save(path)
        return RawSheetCandidate(path, worksheet.title, 1, headers)

    def test_account_6_and_merchandise_weight_are_excluded(self):
        config = load_config(ROOT / "config" / "analysis_rules.yaml")
        with TemporaryDirectory() as temp:
            directory = Path(temp)
            prior = self._make_source(directory, 2025)
            current = self._make_source(directory, 2026)
            manifest = InputManifest(directory / "official.xlsx", {2025: prior, 2026: current}, {})
            datasets = load_cumulative_datasets(manifest, "202607", config)
            current_rows = [row for row in datasets["steel_domestic"].rows if row["year"] == 2026]
            self.assertAlmostEqual(sum(row["amount"] for row in current_rows), 3.0)
            self.assertAlmostEqual(sum(row["weight"] for row in current_rows), 1.0)
            self.assertEqual(CUMULATIVE_REQUIRED_HEADERS - set(prior.headers), set())


class RecipientExclusionTests(unittest.TestCase):
    def _make_source(self, directory: Path, year: int) -> RawSheetCandidate:
        headers = [
            "출고요청년월", "LVL1NM", "LVL2NM", "LVL3NM", "품번2", "계정",
            "중량(Kg)", "원화금액", "달러금액", "거래처", "인수처명",
        ]
        path = directory / f"{year}_steel_export.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(headers)
        worksheet.append([f"{year}07", "스틸로프", "스틸상품", "상품", "CCG", 2, 1_000, 0, 100_000, "HI-LEX", "HI-LEX USA"])
        worksheet.append([f"{year}07", "스틸로프", "SPECIAL ROPE 1", "POWERFLEX", "SR1", 2, 2_000, 0, 200_000, "SCM", "SCM CITRA OY"])
        workbook.save(path)
        return RawSheetCandidate(path, worksheet.title, 1, headers)

    def test_steel_control_cable_is_excluded_only_from_recipient_analysis(self):
        config = load_config(ROOT / "config" / "analysis_rules.yaml")
        config["report_order"] = ["steel_export"]
        with TemporaryDirectory() as temp:
            directory = Path(temp)
            prior = self._make_source(directory, 2025)
            current = self._make_source(directory, 2026)
            manifest = InputManifest(
                directory / "official.xlsx",
                {},
                {("steel_export", 2025): prior, ("steel_export", 2026): current},
            )
            datasets = load_recipient_datasets(manifest, "202607", config)
            rows = datasets["steel_export"].rows
            self.assertEqual({row["primary"] for row in rows}, {"SR1"})
            self.assertEqual({row["secondary"] for row in rows}, {"SR1"})
            self.assertEqual(len(rows), 2)
            self.assertTrue(all(row["weight"] == 2.0 for row in rows))


class RecipientInputRobustnessTests(unittest.TestCase):
    def test_raw_sheet_is_selected_over_one_row_detail_sheet(self):
        headers = [
            "출고요청년월", "LVL1NM", "LVL2NM", "LVL3NM", "계정",
            "원화금액", "달러금액", "거래처", "인수처명",
        ]
        with TemporaryDirectory() as temp:
            path = Path(temp) / "multiple_sheets.xlsx"
            workbook = Workbook()
            detail = workbook.active
            detail.title = "세부 정보1"
            detail.append(["피벗 세부 정보"])
            detail.append([])
            detail.append(headers)
            detail.append(["202606", "스틸로프", "WIRE ROPE", "WRG.I", 2, 130, 0, "거래처", "인수처"])
            raw = workbook.create_sheet("Raw data")
            raw.append(headers)
            raw.append(["202501", "스틸로프", "WIRE ROPE", "WRG.I", 2, 10, 0, "거래처", "인수처"])
            raw.append(["202601", "스틸로프", "WIRE ROPE", "WRG.I", 2, 20, 0, "거래처", "인수처"])
            workbook.save(path)

            candidate = _find_header(path, RECIPIENT_REQUIRED_HEADERS, "출고요청년월", {2025, 2026})
            self.assertIsNotNone(candidate)
            self.assertEqual(candidate.sheet_name, "Raw data")
            self.assertEqual(candidate.header_row, 1)

    def test_report_units_are_not_divided_twice(self):
        metric = OfficialMetric(
            segment="steel_domestic",
            label="제강 내수",
            market="domestic",
            currency_unit="백만원",
            prior_ytd_amount=35,
            current_ytd_amount=100,
            prior_ytd_weight=10,
            current_ytd_weight=20,
        )
        reference = {"steel_domestic": metric}
        self.assertEqual(
            _unit_divisors("steel_domestic", "domestic", 2025, 2026, [35], [10], reference),
            (1.0, 1.0),
        )
        self.assertEqual(
            _unit_divisors("steel_domestic", "domestic", 2025, 2026, [35_000_000], [10_000], reference),
            (1_000_000.0, 1_000.0),
        )

    def test_only_hyundai_samho_alias_is_normalized(self):
        self.assertEqual(_normalize_recipient_name("현대삼호중공업(주)"), "에이치디현대삼호㈜")
        self.assertEqual(_normalize_recipient_name("에이치디현대삼호(주)"), "에이치디현대삼호㈜")
        self.assertEqual(_normalize_recipient_name("한화오션(주)"), "한화오션(주)")


class NarrativeLinkTests(unittest.TestCase):
    def _recipient(self, secondary: str, recipient: str, delta: float) -> dict:
        return {
            "primary": "스텐선재",
            "secondary": secondary,
            "recipient": recipient,
            "amount_delta": delta,
            "previous_amount": 0,
            "current_amount": delta,
            "prior_ytd": 100,
            "current_ytd": 100 + delta,
            "prior_year_exists": "있음",
            "status": "증가",
            "monthly_amounts": {f"2026{month:02d}": 10 + month for month in range(1, 8)},
        }

    def test_secondary_recipient_text_does_not_mix_other_secondary(self):
        config = load_config(ROOT / "config" / "analysis_rules.yaml")
        summary = {"comparison": "전월 대비", "currency_unit": "천달러"}
        secondary = {"primary": "스텐선재", "secondary": "SS SPRING", "amount_delta": 200}
        recipients = [
            self._recipient("SS SPRING", "METALLE SCHMIDT", 150),
            self._recipient("SS C/HEADING", "OTHER CUSTOMER", 500),
        ]
        text = " ".join(_secondary_recipient_paragraphs(summary, secondary, recipients, config, 2026, 7))
        self.assertIn("METALLE SCHMIDT", text)
        self.assertNotIn("OTHER CUSTOMER", text)

    def test_monthly_series_is_expanded_only_for_exception_pattern(self):
        continuous = self._recipient("SS SPRING", "VOGELSANG", 100)
        intermittent = self._recipient("SS SPRING", "METALLE SCHMIDT", 100)
        intermittent["monthly_amounts"] = {"202601": 100, "202603": 120, "202607": 90}
        intermittent["status"] = "출고 재개"
        self.assertNotIn("→", _trend_details(continuous, 2026, 7, "천달러"))
        self.assertIn("→", _trend_details(intermittent, 2026, 7, "천달러"))

    def test_steel_links_level2_directly_to_recipient(self):
        config = load_config(ROOT / "config" / "analysis_rules.yaml")
        summary = {"comparison": "누계", "currency_unit": "백만원"}
        item = {
            "primary": "WIRE ROPE",
            "previous_amount": 1_000,
            "current_amount": 1_800,
            "amount_delta": 800,
            "amount_pct": 0.8,
            "previous_weight": 500,
            "current_weight": 900,
            "weight_delta": 400,
            "weight_pct": 0.8,
            "previous_price": 2.0,
            "current_price": 2.0,
            "price_pct": 0.0,
        }
        recipients = [{
            "primary": "WIRE ROPE",
            "secondary": "WIRE ROPE",
            "recipient": "동성기업 주식회사",
            "amount_delta": 521,
            "previous_amount": 0,
            "current_amount": 521,
            "prior_ytd": 0,
            "current_ytd": 521,
            "prior_year_exists": "없음",
            "status": "당해 신규",
            "monthly_amounts": {f"2026{month:02d}": 70 for month in range(1, 8)},
        }]
        secondary_items = [{"primary": "WIRE ROPE", "secondary": "WRG.I", "amount_delta": 792}]
        text = " ".join(_focus_paragraphs(summary, item, secondary_items, recipients, config, "steel_domestic", 2026, 7))
        self.assertIn("WIRE ROPE 증가분은 동성기업 주식회사향", text)
        self.assertNotIn("WRG.I", text)
        self.assertNotIn("하위 분류", text)

    def test_recipient_mix_shift_surfaces_offsetting_movements_and_price_mix(self):
        config = load_config(ROOT / "config" / "analysis_rules.yaml")
        product = {"previous_amount": 500, "current_amount": 539, "amount_delta": 39}

        def compared(name: str, previous_amount: float, current_amount: float, previous_weight: float, current_weight: float) -> dict:
            previous_price = safe_unit_price(previous_amount, previous_weight)
            current_price = safe_unit_price(current_amount, current_weight)
            return {
                "key": ("스텐선재", "SS GENERAL", name),
                "previous_amount": previous_amount,
                "current_amount": current_amount,
                "amount_delta": current_amount - previous_amount,
                "amount_pct": pct_change(current_amount, previous_amount),
                "previous_weight": previous_weight,
                "current_weight": current_weight,
                "weight_delta": current_weight - previous_weight,
                "weight_pct": pct_change(current_weight, previous_weight),
                "previous_price": previous_price,
                "current_price": current_price,
                "price_pct": pct_change(current_price, previous_price) if current_price is not None and previous_price is not None else None,
                "previous_rows": 1,
                "current_rows": 1,
            }

        items = [
            compared("일동", 278, 184, 11.0, 12.5),
            compared("동아", 158, 0, 8.0, 0.0),
            compared("동방", 0, 101, 0.0, 5.0),
            compared("비앤", 0, 100, 0.0, 4.0),
            compared("와이어", 0, 90, 0.0, 3.0),
        ]
        names = {
            ("스텐선재", "SS GENERAL", name): name
            for name in ("일동", "동아", "동방", "비앤", "와이어")
        }
        shift = _detect_recipient_mix_shift(
            "stainless_domestic",
            "스텐 내수",
            "전월 대비",
            "스텐선재",
            "SS GENERAL",
            product,
            items,
            names,
            config,
        )
        self.assertIsNotNone(shift)
        summary = {"currency_unit": "백만원"}
        text = " ".join(_recipient_mix_shift_paragraphs(summary, [shift], set()))
        self.assertIn("SS GENERAL은 39백만원 증가했지만", text)
        self.assertIn("일동향 94백만원", text)
        self.assertIn("278에서 184백만원", text)
        self.assertIn("11.0에서 12.5톤", text)
        self.assertIn("평균단가는", text)
        self.assertIsNone(
            _detect_recipient_mix_shift(
                "stainless_domestic",
                "스텐 내수",
                "누계",
                "스텐선재",
                "SS GENERAL",
                product,
                items,
                names,
                config,
            )
        )

    def test_recipient_mix_shift_ignores_one_sided_change(self):
        config = load_config(ROOT / "config" / "analysis_rules.yaml")
        item = {
            "key": ("스텐선재", "SS GENERAL", "동방"),
            "previous_amount": 0,
            "current_amount": 300,
            "amount_delta": 300,
            "amount_pct": None,
            "previous_weight": 0,
            "current_weight": 10,
            "weight_delta": 10,
            "weight_pct": None,
            "previous_price": None,
            "current_price": 30,
            "price_pct": None,
            "previous_rows": 0,
            "current_rows": 1,
        }
        shift = _detect_recipient_mix_shift(
            "stainless_domestic",
            "스텐 내수",
            "전월 대비",
            "스텐선재",
            "SS GENERAL",
            {"previous_amount": 0, "current_amount": 300, "amount_delta": 300},
            [item],
            {("스텐선재", "SS GENERAL", "동방"): "동방"},
            config,
        )
        self.assertIsNone(shift)


if __name__ == "__main__":
    unittest.main()
