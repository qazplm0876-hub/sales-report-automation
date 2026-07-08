"""
월별 사업계획대비 매출실적 분석 자동화 스크립트 (통합본)
============================================================
작성일: 2026-05

[사용법]
1. MONTH, 파일 경로 설정
2. python3 monthly_report_final.py

[입력 파일]
  - 템플릿:       사업계획대비_매출실적분석_26년_{N}월_.xlsx
  - 매출계획:     월별_사업계획.xlsx
  - 판매량계획:   제품판매량_사업계획.xlsx
  - 해외계획:     2_해외현지내수판매계획_청도_비나_DCV_.xlsx
  - 당해 raw:     2026{N}월누계.xls
  - 전년 raw:     2025{N}월누계.xls
  - DSR 손익:     DSR_{N:02d}_상원.xlsx
  - 제강 손익:    제강{N:02d}_상원.xlsx

[자동화 범위]
  섹션1. 전체매출실적     C(전월)/D(당월계획)/E(당월실적)/H(전년실적)/I(당해계획)/J(당해실적)
  섹션2. 제품판매량       C/D/E/H/I/J
  섹션3. 제품판매단가     C(전전월)/D(전월)/E(당월실적)/H(전년실적)/I(당해계획)/J(당해실적)
  섹션4. 통합손익요약     C(전월)/D(당월실적)/F(전년실적)/G(당해실적) ← 이익율은 수식 자동계산

[수동 입력 항목]
  - 섹션1 기타매출(E/H/J열)
  - 섹션1 C1(환율)
  - 섹션4 합섬/스텐 당해계획(H열)
"""

from openpyxl import load_workbook
import pandas as pd
import shutil

# =============================================
# ★ 매월 여기만 수정 ★
# =============================================
MONTH = 5   # 당월 (4=4월, 5=5월, ...)
PREV_MONTH = f"2026{MONTH-1:02d}"   # 전월 요청월 키 (예: 202603)
CURR_MONTH = f"2026{MONTH:02d}"     # 당월 요청월 키 (예: 202604)
PREV2_MONTH = f"2026{MONTH-2:02d}"  # 전전월 요청월 키 (예: 202602)

# 파일 경로
TEMPLATE_FILE  = f"사업계획대비_매출실적분석_26년_{MONTH}월_.xlsx"
PLAN_FILE      = "월별_사업계획.xlsx"
VOL_FILE       = "제품판매량_사업계획.xlsx"
OVERSEAS_FILE  = "2_해외현지내수판매계획_청도_비나_DCV_.xlsx"
import os
def _find_raw(pattern_base):
    for ext in ['.xls', '.xlsx']:
        if os.path.exists(pattern_base + ext):
            return pattern_base + ext
    raise FileNotFoundError(f'{pattern_base}.xls 또는 .xlsx 파일을 찾을 수 없습니다.')

def _customer_file(name):
    candidates = [
        os.path.join(os.getcwd(), name),
        os.path.join(os.path.expanduser('~'), 'Downloads', name),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    raise FileNotFoundError(f'{name} 거래처 상세 파일을 찾을 수 없습니다.')

RAW_26_FILE    = _find_raw(f'2026{MONTH}월누계')
RAW_25_FILE    = _find_raw(f'2025{MONTH}월누계')
DSR_FILE       = f"DSR_{MONTH:02d}_상원.xlsx"
STEEL_FILE     = f"제강{MONTH:02d}_상원.xlsx"
OUTPUT_FILE    = f"사업계획대비_매출실적분석_26년_{MONTH}월_완성_CLAUDE.xlsx"
CUSTOMER_FILES = {
    '합섬': _customer_file('1.xlsx'),
    '스텐': _customer_file('2.xlsx'),
    '제강': _customer_file('3.xlsx'),
}


# =============================================
# 공통 유틸
# =============================================
def plan_month_row(month):
    """월별_사업계획.xlsx 행번호: 1월=6,2월=7,3월=8,(1Q=9),4월=10,..."""
    return 5 + month + (month-1)//3

def plan_quarter_row(month):
    """직전 분기 누계행"""
    quarter = (month-1)//3
    if quarter == 0: return None
    return 5 + (quarter*3) + (quarter-1) + 1

def overseas_month_row(month):
    """해외계획 파일 월별 행"""
    return 5 + month + (month-1)//3

def vol_col_kg(month):  return 4 + month*2   # 제품판매량 중량 열
def vol_col_amt(month): return 5 + month*2   # 제품판매량 금액 열

def ytd_kg(ws, row, month):
    return sum(ws.cell(row, vol_col_kg(m)).value or 0 for m in range(1, month+1))

def ytd_amt(ws, row, month):
    return sum(ws.cell(row, vol_col_amt(m)).value or 0 for m in range(1, month+1))

def unit_price_vol(ws, rows, month):
    """제품판매량 파일 기준 단가 = 누적금액/누적중량"""
    kg  = sum(ytd_kg(ws, r, month) for r in rows)
    amt = sum(ytd_amt(ws, r, month) for r in rows)
    return amt/kg if kg else 0

def load_raw(filepath):
    """raw xls/xlsx 파일 로드 및 전처리"""
    if filepath.endswith('.xlsx'):
        df = pd.read_excel(filepath, engine='openpyxl')
    else:
        df = pd.read_excel(filepath, engine='xlrd')
    # 회사/계정/요청월 강제 문자열 변환 (xlsx에서 정수로 읽히는 문제 해결)
    df['회사']    = df['회사'].astype(str).str.strip()
    df['계정']    = df['계정'].astype(str).str.strip()
    df['요청월']  = df['요청월'].astype(str).str.strip()
    df = df[df['회사'] != '6']
    df = df[df['계정'] != '6']
    df = df.drop(columns=['원화금액'])
    df['중량']        = pd.to_numeric(df['중량'],        errors='coerce') / 1000
    df['달러금액']    = pd.to_numeric(df['달러금액'],    errors='coerce') / 1000
    df['한국원화금액'] = pd.to_numeric(df['한국원화금액'], errors='coerce') / 1_000_000
    return df

def raw_kg(df, 부문, 내수수출, lvl1, lvl2=None):
    mask = (df['부문']==부문) & (df['내수/수출'].isin(내수수출 if isinstance(내수수출,list) else [내수수출]))
    mask &= df['레벨1명'].isin(lvl1 if isinstance(lvl1,list) else [lvl1])
    if lvl2: mask &= df['레벨2명'].isin(lvl2 if isinstance(lvl2,list) else [lvl2])
    return df[mask]['중량'].sum()

def raw_unit(df, 부문, 내수수출, lvl1, lvl2=None, amt='달러금액'):
    mask = (df['부문']==부문) & (df['내수/수출'].isin(내수수출 if isinstance(내수수출,list) else [내수수출]))
    mask &= df['레벨1명'].isin(lvl1 if isinstance(lvl1,list) else [lvl1])
    if lvl2: mask &= df['레벨2명'].isin(lvl2 if isinstance(lvl2,list) else [lvl2])
    sub = df[mask]
    kg = sub['중량'].sum(); a = sub[amt].sum()
    return a/kg if kg else 0

def raw_sales(df, 부문, 내수수출, amt='한국원화금액'):
    mask = (df['부문']==부문) & (df['내수/수출'].isin(내수수출 if isinstance(내수수출,list) else [내수수출]))
    return df[mask][amt].sum()


# =============================================
# 섹션1: 전체매출실적
# 채우는 열: C(전월)=3, D(당월계획)=4, E(당월실적)=5, H(전년실적)=8, I(당해계획)=9, J(당해실적)=10
# =============================================
def fill_section1(ws4, wsp, df26, df25, month):
    mr = plan_month_row(month)
    qr = plan_quarter_row(month)

    def vp(row, col): return wsp.cell(row=row, column=col).value or 0
    def s4(row, col, val): ws4.cell(row=row, column=col).value = val

    def cum_plan(base_mr, col):
        """직전분기누계 + 분기내 당월까지 합산
        base_mr은 해당 부문의 당월 행. 부문 오프셋(base_mr - plan_month_row(month))을 이용해
        같은 부문의 다른 월/분기 행을 찾는다."""
        offset = base_mr - plan_month_row(month)
        local_qr = qr + offset if qr else None
        q = (month-1)//3; q_start = q*3+1
        total = (vp(local_qr, col) if local_qr else 0)
        for m in range(q_start, month+1):
            total += vp(plan_month_row(m) + offset, col) or 0
        return total

    ST=23; JG=46; CD=69; VN=92  # 계획파일 섹션별 행 오프셋

    VINA = ['DSR VINA-현지내수','CABLE VINA-내수','CABLE VINA-수출']

    # 당월계획(D) / 당해계획누계(I)
    s4(7,4,vp(mr,4));    s4(7,9,cum_plan(mr,4))     # 합섬수출$
    s4(8,4,vp(mr+ST,4)); s4(8,9,cum_plan(mr+ST,4))  # 스텐수출$
    s4(9,4,vp(mr+JG,4)); s4(9,9,cum_plan(mr+JG,4))  # 제강수출$
    s4(16,4,vp(mr,7));   s4(16,9,cum_plan(mr,7))     # 합섬내수
    s4(17,4,vp(mr+ST,7));s4(17,9,cum_plan(mr+ST,7)) # 스텐내수
    s4(18,4,vp(mr+JG,7));s4(18,9,cum_plan(mr+JG,7)) # 제강내수
    s4(24,4,vp(mr+CD,5)); s4(25,4,vp(mr+CD,8))       # 청도 합섬/제강
    s4(27,4,vp(mr+VN,5)); s4(28,4,vp(mr+VN,8))       # 비나 합섬/제강
    if ws4.cell(27,9).value is None: s4(27,9,'=+D27*$C$5')
    if ws4.cell(28,9).value is None: s4(28,9,'=+D28*$C$5')
    # 부산물
    bp = [wsp.cell(r,5).value or 0 for r in range(119,127)]
    s4(33,4,bp[0]+bp[1]+bp[2]); s4(34,4,bp[3]+bp[4]); s4(35,4,bp[5]+bp[6]+bp[7])

    # 전월(C) / 당월실적(E) / 당해실적(J) - raw data
    for col, df in [(3, df26[df26['요청월']==PREV_MONTH]),
                    (5, df26[df26['요청월']==CURR_MONTH]),
                    (10, df26)]:
        s4(7,col,  raw_sales(df,'합섬','수출','달러금액'))
        s4(8,col,  raw_sales(df,'STS', '수출','달러금액'))
        s4(9,col,  raw_sales(df,'제강','수출','달러금액'))
        s4(11,col, raw_sales(df,'합섬','수출','한국원화금액'))
        s4(12,col, raw_sales(df,'STS', '수출','한국원화금액'))
        s4(13,col, raw_sales(df,'제강','수출','한국원화금액'))
        s4(16,col, raw_sales(df,'합섬','내수','한국원화금액'))
        s4(17,col, raw_sales(df,'STS', '내수','한국원화금액'))
        s4(18,col, raw_sales(df,'제강','내수','한국원화금액'))
        s4(24,col, raw_sales(df,'합섬','청도DSR-현지내수','한국원화금액'))
        s4(25,col, raw_sales(df,'제강','청도DSR-현지내수','한국원화금액'))
        s4(27,col, sum(raw_sales(df,'합섬',k,'한국원화금액') for k in VINA))
        s4(28,col, sum(raw_sales(df,'제강',k,'한국원화금액') for k in VINA))

    # 전년실적(H) - 2025 raw
    for col, df in [(8, df25)]:
        s4(7,col,  raw_sales(df,'합섬','수출','달러금액'))
        s4(8,col,  raw_sales(df,'STS', '수출','달러금액'))
        s4(9,col,  raw_sales(df,'제강','수출','달러금액'))
        s4(11,col, raw_sales(df,'합섬','수출','한국원화금액'))
        s4(12,col, raw_sales(df,'STS', '수출','한국원화금액'))
        s4(13,col, raw_sales(df,'제강','수출','한국원화금액'))
        s4(16,col, raw_sales(df,'합섬','내수','한국원화금액'))
        s4(17,col, raw_sales(df,'STS', '내수','한국원화금액'))
        s4(18,col, raw_sales(df,'제강','내수','한국원화금액'))
        s4(24,col, raw_sales(df,'합섬','청도DSR-현지내수','한국원화금액'))
        s4(25,col, raw_sales(df,'제강','청도DSR-현지내수','한국원화금액'))
        s4(27,col, sum(raw_sales(df,'합섬',k,'한국원화금액') for k in VINA))
        s4(28,col, sum(raw_sales(df,'제강',k,'한국원화금액') for k in VINA))

    print(f"[완료] 섹션1")


# =============================================
# 섹션2: 제품판매량
# =============================================
def fill_section2(ws4, ws_hs, ws_st, ws_jg, ws_cd, ws_vn, ws_dc, df26, df25, month):
    m = month
    VINA = ['DSR VINA-현지내수','CABLE VINA-내수','CABLE VINA-수출']

    HS_LVL1     = ['합섬기타','합섬방사','합섬비방사','합섬웨빙','합섬특수']
    ST_LVL1     = ['스텐로프','스텐선재']
    JG_LVL1     = ['스틸로프','스틸선재']
    JG_LVL1_CD  = ['스틸로프','스틸선재','스틸상품']
    GS_LVL2     = ['GUY STRAND']

    WR   = ['WIRE ROPE','SPECIAL ROPE 1','SPECIAL ROPE 2']
    CC   = ['CONTROL CABLE']
    GS   = ['GUY STRAND']
    WIRE = ['WIRE']
    OT   = ['OT WIRE']
    IT   = ['IT WIRE']

    def s4(row, col, val):
        if isinstance(val, str) and val.startswith('='):
            ws4.cell(row=row, column=col).value = val
        else:
            ws4.cell(row=row, column=col).value = round(float(val),3) if val else 0

    # 당월계획(D) / 당해계획(I) - 판매량 계획파일
    s4(48,4,ws_hs.cell(66,vol_col_kg(m)).value or 0)
    s4(48,9,ytd_kg(ws_hs,66,m))
    s4(49,4,ws_st.cell(49,vol_col_kg(m)).value or 0)
    s4(49,9,ytd_kg(ws_st,49,m))
    s4(50,4,'=SUM(D51:D56)'); s4(50,9,'=SUM(I51:I56)')
    for row,rows_exp,rows_all in [
        (51,[46,47,48],[46,47,48]),(52,[49],[49]),(53,[51],[51]),
        (54,[52],[52]),(55,[53],[53]),(56,[54],[54])]:
        s4(row,4,sum(ws_jg.cell(r,vol_col_kg(m)).value or 0 for r in rows_exp))
        s4(row,9,sum(ytd_kg(ws_jg,r,m) for r in rows_all))
    s4(58,4,ws_hs.cell(43,vol_col_kg(m)).value or 0)
    s4(58,9,ytd_kg(ws_hs,43,m))
    s4(59,4,ws_st.cell(33,vol_col_kg(m)).value or 0)
    s4(59,9,ytd_kg(ws_st,33,m))
    s4(60,4,'=SUM(D61:D66)'); s4(60,9,'=SUM(I61:I66)')
    for row,rows in [(61,[26,27,28]),(62,[29]),(63,[31]),(64,[32]),(65,[33]),(66,[34])]:
        s4(row,4,sum(ws_jg.cell(r,vol_col_kg(m)).value or 0 for r in rows))
        s4(row,9,sum(ytd_kg(ws_jg,r,m) for r in rows))
    ov_row = overseas_month_row(m)
    s4(68,4,ws_cd.cell(ov_row,3).value); s4(68,9,'=D68*$C$46')
    s4(69,4,ws_cd.cell(ov_row,5).value); s4(69,9,'=D69*$C$46')
    s4(71,4,ws_vn.cell(ov_row,3).value); s4(71,9,'=D71*$C$46')
    s4(72,4,ws_dc.cell(ov_row,3).value); s4(72,9,'=D72*$C$46')

    # C(전월)/E(당월)/H(전년)/J(당해) - raw data
    for col, df in [(3, df26[df26['요청월']==PREV_MONTH]),
                    (5, df26[df26['요청월']==CURR_MONTH]),
                    (8, df25),
                    (10, df26)]:
        s4(48,col,raw_kg(df,'합섬','수출',HS_LVL1))
        s4(49,col,raw_kg(df,'STS', '수출',ST_LVL1))
        s4(51,col,raw_kg(df,'제강','수출',JG_LVL1,WR))
        s4(52,col,raw_kg(df,'제강','수출',JG_LVL1,CC))
        s4(53,col,raw_kg(df,'제강',['수출','내수'],['스틸ＳＴ'],GS))  # GS 수출+내수 합산
        s4(54,col,raw_kg(df,'제강','수출',JG_LVL1,WIRE))
        s4(55,col,raw_kg(df,'제강','수출',JG_LVL1,OT))
        s4(56,col,raw_kg(df,'제강','수출',JG_LVL1,IT))
        s4(58,col,raw_kg(df,'합섬','내수',HS_LVL1))
        s4(59,col,raw_kg(df,'STS', '내수',ST_LVL1))
        s4(61,col,raw_kg(df,'제강','내수',JG_LVL1,WR))
        s4(62,col,raw_kg(df,'제강','내수',JG_LVL1,CC))
        s4(63,col,raw_kg(df,'제강','내수',JG_LVL1,GS) + raw_kg(df,'제강','내수',['스틸ＳＴ'],GS))
        s4(64,col,raw_kg(df,'제강','내수',JG_LVL1,WIRE))
        s4(65,col,raw_kg(df,'제강','내수',JG_LVL1,OT))
        s4(66,col,raw_kg(df,'제강','내수',JG_LVL1,IT))
        s4(68,col,raw_kg(df,'합섬','청도DSR-현지내수',HS_LVL1))
        s4(69,col,raw_kg(df,'제강','청도DSR-현지내수',JG_LVL1_CD))
        s4(71,col,raw_kg(df,'합섬',VINA,HS_LVL1))
        s4(72,col,raw_kg(df,'제강',VINA,JG_LVL1))

    print(f"[완료] 섹션2")


# =============================================
# 섹션3: 제품판매단가
# =============================================
def fill_section3(ws4, ws_hs, ws_st, ws_jg, wbov, df26, df25, month):
    m = month
    VINA = ['DSR VINA-현지내수','CABLE VINA-내수','CABLE VINA-수출']
    HS_LVL1_CD = ['합섬기타','합섬방사','합섬비방사','합섬웨빙','합섬특수']
    JG_LVL1_CD = ['스틸로프','스틸상품']
    JG_LVL1_VN = ['스틸로프','스틸선재']

    def s4(row, col, val): ws4.cell(row=row, column=col).value = round(float(val),6) if val else 0
    def up(ws, rows): return unit_price_vol(ws, rows, m)

    # 당해계획(I) - 판매량계획파일 기준 단가
    s4(83,9,up(ws_hs,[52,53,54,55])); s4(84,9,up(ws_hs,[57,58,59]))
    s4(85,9,up(ws_hs,[61]));          s4(86,9,up(ws_hs,[62]))
    s4(87,9,up(ws_hs,[29,30,31,32])); s4(88,9,up(ws_hs,[34,35,36]))
    s4(89,9,up(ws_hs,[38]));          s4(90,9,up(ws_hs,[39]))
    s4(91,9,up(ws_st,[38,39]));       s4(92,9,up(ws_st,[41,42,43,44,45,46,47]))
    s4(93,9,up(ws_st,[22,23]));       s4(94,9,up(ws_st,[25,26,27,28,29,30,31]))
    s4(95,9,up(ws_jg,[46,47,48]));    s4(96,9,up(ws_jg,[49]))
    s4(97,9,up(ws_jg,[51]));          s4(98,9,up(ws_jg,[52]))
    s4(99,9,up(ws_jg,[53]));          s4(100,9,up(ws_jg,[54]))
    s4(101,9,up(ws_jg,[26,27,28]));   s4(102,9,up(ws_jg,[29]))
    s4(103,9,up(ws_jg,[31]));         s4(104,9,up(ws_jg,[32]))
    s4(105,9,up(ws_jg,[33]));         s4(106,9,up(ws_jg,[34]))

    # C(전전월)/D(전월)/E(당월)/H(전년)/J(당해) - raw data 기준 단가
    PREV2 = df26[df26['요청월']==PREV2_MONTH]
    PREV  = df26[df26['요청월']==PREV_MONTH]
    CURR  = df26[df26['요청월']==CURR_MONTH]

    mappings = [
        # (행, 부문, 내수수출, lvl1, lvl2, 금액열)
        (83,'합섬','수출',['합섬방사'],None,'달러금액'),
        (84,'합섬','수출',['합섬비방사'],None,'달러금액'),
        (85,'합섬','수출',['합섬웨빙'],None,'달러금액'),
        (86,'합섬','수출',['합섬특수'],None,'달러금액'),
        (87,'합섬','내수',['합섬방사'],None,'한국원화금액'),
        (88,'합섬','내수',['합섬비방사'],None,'한국원화금액'),
        (89,'합섬','내수',['합섬웨빙'],None,'한국원화금액'),
        (90,'합섬','내수',['합섬특수'],None,'한국원화금액'),
        (91,'STS','수출',['스텐로프'],None,'달러금액'),
        (92,'STS','수출',['스텐선재'],None,'달러금액'),
        (93,'STS','내수',['스텐로프'],None,'한국원화금액'),
        (94,'STS','내수',['스텐선재'],None,'한국원화금액'),
        (95,'제강','수출',['스틸로프'],['WIRE ROPE','SPECIAL ROPE 1','SPECIAL ROPE 2'],'달러금액'),
        (96,'제강','수출',['스틸로프'],['CONTROL CABLE'],'달러금액'),
        (97,'제강','수출',['스틸ＳＴ'],['GUY STRAND'],'달러금액'),
        (98,'제강','수출',['스틸선재'],['WIRE'],'달러금액'),
        (99,'제강','수출',['스틸선재'],['OT WIRE'],'달러금액'),
        (100,'제강','수출',['스틸선재'],['IT WIRE'],'달러금액'),
        (101,'제강','내수',['스틸로프'],['WIRE ROPE','SPECIAL ROPE 1','SPECIAL ROPE 2'],'한국원화금액'),
        (102,'제강','내수',['스틸로프'],['CONTROL CABLE'],'한국원화금액'),
        (103,'제강','내수',['스틸ＳＴ'],['GUY STRAND'],'한국원화금액'),
        (104,'제강','내수',['스틸선재'],['WIRE'],'한국원화금액'),
        (105,'제강','내수',['스틸선재'],['OT WIRE'],'한국원화금액'),
        (106,'제강','내수',['스틸선재'],['IT WIRE'],'한국원화금액'),
    ]
    for (row, 부문, 내수수출, lvl1, lvl2, amt) in mappings:
        s4(row,3, raw_unit(PREV2, 부문, 내수수출, lvl1, lvl2, amt))
        s4(row,4, raw_unit(PREV,  부문, 내수수출, lvl1, lvl2, amt))
        s4(row,5, raw_unit(CURR,  부문, 내수수출, lvl1, lvl2, amt))
        s4(row,8, raw_unit(df25,  부문, 내수수출, lvl1, lvl2, amt))
        s4(row,10,raw_unit(df26,  부문, 내수수출, lvl1, lvl2, amt))

    # 청도/비나 단가 (I열은 해외계획파일 환율변환, C/D/E/H/J는 raw)
    ws_cd_ov = wbov['청도26']; ws_vn_ov = wbov['비나26']; ws_dc_ov = wbov['DCV26']

    def ov_price_plan(ws, kg_col, amt_col, rate):
        def ov_row(mo): return 5 + mo + (mo-1)//3
        def ytd_s(col): return sum(ws.cell(ov_row(mo),col).value or 0 for mo in range(1,m+1))
        kg = ytd_s(kg_col); amt = ytd_s(amt_col)
        return (amt * rate / 1_000_000) / kg if kg else 0

    s4(107,9,ov_price_plan(ws_cd_ov,3,4,200))   # 청도합섬 계획단가
    s4(108,9,ov_price_plan(ws_cd_ov,5,6,200))   # 청도제강 계획단가
    s4(109,9,ov_price_plan(ws_vn_ov,3,4,0.055)) # 비나합섬 계획단가
    s4(110,9,ov_price_plan(ws_dc_ov,3,4,0.055)) # 비나제강 계획단가

    for col, df in [(3,PREV2),(4,PREV),(5,CURR),(8,df25),(10,df26)]:
        s4(107,col,raw_unit(df,'합섬','청도DSR-현지내수',HS_LVL1_CD,None,'한국원화금액'))
        s4(108,col,raw_unit(df,'제강','청도DSR-현지내수',JG_LVL1_CD,None,'한국원화금액'))
        s4(109,col,raw_unit(df,'합섬',VINA,HS_LVL1_CD,None,'한국원화금액'))
        s4(110,col,raw_unit(df,'제강',VINA,JG_LVL1_VN,None,'한국원화금액'))

    print(f"[완료] 섹션3")


# =============================================
# 섹션4: 통합손익요약
# 채우는 열: C(전월)=3, D(당월실적)=4, F(전년실적)=6, G(당해실적)=7
# 이익율 E,H열은 수식으로 자동계산
# =============================================
def fill_section4(ws4, wb_dsr, wb_steel, month):
    m = month

    def s4(row, col, val): ws4.cell(row=row, column=col).value = val
    def mv(v): return round((v or 0)/1000, 1)
    def tot(*vals): return round(sum(vals), 1)

    # ── 합섬 ────────────────────────────────
    ws_hk = wb_dsr['합섬(한국)']
    ws_hc = wb_dsr['합섬(중국)']
    ws_hv = wb_dsr['합섬(베트남)']

    # 합섬 열: 1월=3,2월=4,...  전월=m+1, 당월=m+2, 25년1~m월 누계
    def hs_col_prev(mo): return mo+1   # 26.{m-1}월
    def hs_col_curr(mo): return mo+2   # 26.{m}월
    def hs_ytd25(ws, row):
        return sum(ws.cell(row, 3+i).value or 0 for i in range(m))  # 25.01~25.m
    def hs_ytd26_col(mo): return 5 + mo + 22  # 상반기누계=AA=27(4월기준), 월마다 변동

    # 합섬 열 위치는 월마다 다르므로 헤더에서 찾기
    def find_col(ws, target):
        for c in range(1, 100):
            if str(ws.cell(3,c).value) == str(target): return c
        return None

    hk_prev = find_col(ws_hk, f'26.{m-1:02d}') or (m+1)
    hk_curr = find_col(ws_hk, f'26.{m:02d}')   or (m+2)
    hk_ytd  = find_col(ws_hk, '상반기') or 27

    hc_prev = find_col(ws_hc, f'26.{m-1:02d}') or (m+1)
    hc_curr = find_col(ws_hc, f'26.{m:02d}')   or (m+2)
    hc_ytd  = find_col(ws_hc, '상반기') or 27

    hv_prev = find_col(ws_hv, f'26.{m-1:02d}') or (m+1)
    hv_curr = find_col(ws_hv, f'26.{m:02d}')   or (m+2)
    hv_ytd  = find_col(ws_hv, '상반기') or 27

    def hs_ytd25_hk(row): return sum(ws_hk.cell(row,c).value or 0 for c in range(hk_curr-m, hk_curr))
    def hs_ytd25_hc(row): return sum(ws_hc.cell(row,c).value or 0 for c in range(hc_curr-m, hc_curr))
    def hs_ytd25_hv(row): return sum(ws_hv.cell(row,c).value or 0 for c in range(hv_curr-m, hv_curr))

    hk = {'매출': [mv(ws_hk.cell(4,hk_prev).value), mv(ws_hk.cell(4,hk_curr).value),
                   mv(hs_ytd25_hk(4)), mv(ws_hk.cell(4,hk_ytd).value)],
          '이익': [mv(ws_hk.cell(9,hk_prev).value), mv(ws_hk.cell(9,hk_curr).value),
                   mv(hs_ytd25_hk(9)), mv(ws_hk.cell(9,hk_ytd).value)]}
    hc = {'매출': [mv(ws_hc.cell(4,hc_prev).value), mv(ws_hc.cell(4,hc_curr).value),
                   mv(hs_ytd25_hc(4)), mv(ws_hc.cell(4,hc_ytd).value)],
          '이익': [mv(ws_hc.cell(9,hc_prev).value), mv(ws_hc.cell(9,hc_curr).value),
                   mv(hs_ytd25_hc(9)), mv(ws_hc.cell(9,hc_ytd).value)]}
    hv = {'매출': [mv(ws_hv.cell(4,hv_prev).value), mv(ws_hv.cell(4,hv_curr).value),
                   mv(hs_ytd25_hv(4)), mv(ws_hv.cell(4,hv_ytd).value)],
          '이익': [mv(ws_hv.cell(9,hv_prev).value), mv(ws_hv.cell(9,hv_curr).value),
                   mv(hs_ytd25_hv(9)), mv(ws_hv.cell(9,hv_ytd).value)]}
    ht = {'매출': [tot(hk['매출'][i],hc['매출'][i],hv['매출'][i]) for i in range(4)],
          '이익': [tot(hk['이익'][i],hc['이익'][i],hv['이익'][i]) for i in range(4)]}

    # ── 스텐 ────────────────────────────────
    ws_sk = wb_dsr['STS(한국)']
    ws_sv = wb_dsr['STS(베트남)']
    sk_prev = find_col(ws_sk, f'26.{m-1:02d}')
    sk_curr = find_col(ws_sk, f'26.{m:02d}')
    sk_ytd  = find_col(ws_sk, '상반기')
    sv_prev = find_col(ws_sv, f'26.{m-1:02d}')
    sv_curr = find_col(ws_sv, f'26.{m:02d}')
    sv_ytd  = find_col(ws_sv, '상반기')

    def st_ytd25_sk(row): return sum(ws_sk.cell(row,c).value or 0 for c in range(sk_curr-m, sk_curr))
    def st_ytd25_sv(row): return sum(ws_sv.cell(row,c).value or 0 for c in range(sv_curr-m, sv_curr))

    sk = {'매출': [mv(ws_sk.cell(4,sk_prev).value), mv(ws_sk.cell(4,sk_curr).value),
                   mv(st_ytd25_sk(4)), mv(ws_sk.cell(4,sk_ytd).value)],
          '이익': [mv(ws_sk.cell(9,sk_prev).value), mv(ws_sk.cell(9,sk_curr).value),
                   mv(st_ytd25_sk(9)), mv(ws_sk.cell(9,sk_ytd).value)]}
    sv = {'매출': [mv(ws_sv.cell(4,sv_prev).value), mv(ws_sv.cell(4,sv_curr).value),
                   mv(st_ytd25_sv(4)), mv(ws_sv.cell(4,sv_ytd).value)],
          '이익': [mv(ws_sv.cell(9,sv_prev).value), mv(ws_sv.cell(9,sv_curr).value),
                   mv(st_ytd25_sv(9)), mv(ws_sv.cell(9,sv_ytd).value)]}
    st = {'매출': [tot(sk['매출'][i],sv['매출'][i]) for i in range(4)],
          '이익': [tot(sk['이익'][i],sv['이익'][i]) for i in range(4)]}

    # ── 제강 ────────────────────────────────
    ws_jg_pl = wb_steel['포괄손익']
    def jg(row,col): return ws_jg_pl.cell(row,col).value or 0
    jg_prev = find_col(ws_jg_pl, f'2026-{m-1:02d}') or 47
    jg_curr = find_col(ws_jg_pl, f'2026-{m:02d}')   or 48
    def jg_ytd25(row): return sum(jg(row, jg_curr-m+i) for i in range(m))  # 25년 1~m월

    jt = {'매출': [mv(sum(jg(r,jg_prev) for r in [5,9,13,17])),
                   mv(sum(jg(r,jg_curr) for r in [5,9,13,17])),
                   mv(jg_ytd25(5)+jg_ytd25(9)+jg_ytd25(13)+jg_ytd25(17)),
                   mv(sum(jg(r,6) for r in [5,9,13,17]))],
          '이익': [mv(jg(22,jg_prev)), mv(jg(22,jg_curr)), mv(jg_ytd25(22)), mv(jg(22,6))]}
    jk = {'매출': [mv(jg(5,jg_prev)), mv(jg(5,jg_curr)), mv(jg_ytd25(5)), mv(jg(5,6))],
          '이익': [mv(jg(24,jg_prev)), mv(jg(24,jg_curr)), mv(jg_ytd25(24)), mv(jg(24,6))]}
    jc = {'매출': [mv(jg(9,jg_prev)), mv(jg(9,jg_curr)), mv(jg_ytd25(9)), mv(jg(9,6))],
          '이익': [mv(jg(26,jg_prev)), mv(jg(26,jg_curr)), mv(jg_ytd25(26)), mv(jg(26,6))]}
    jv = {'매출': [mv(jg(13,jg_prev)), mv(jg(13,jg_curr)), mv(jg_ytd25(13)), mv(jg(13,6))],
          '이익': [mv(jg(32,jg_prev)), mv(jg(32,jg_curr)), mv(jg_ytd25(32)), mv(jg(32,6))]}
    jd = {'매출': [mv(jg(17,jg_prev)), mv(jg(17,jg_curr)), mv(jg_ytd25(17)), mv(jg(17,6))],
          '이익': [mv(jg(38,jg_prev)), mv(jg(38,jg_curr)), mv(jg_ytd25(38)), mv(jg(38,6))]}

    COLS = [3, 4, 6, 7]
    def fill(매출행, 이익행, data):
        for i, col in enumerate(COLS):
            s4(매출행, col, data['매출'][i])
            s4(이익행, col, data['이익'][i])

    fill(116,117,ht); fill(119,120,hk); fill(122,123,hc); fill(125,126,hv)
    fill(128,129,st); fill(131,132,sk); fill(134,135,sv)
    fill(137,138,jt); fill(140,141,jk); fill(143,144,jc)
    fill(146,147,jv); fill(149,150,jd)

    print(f"[완료] 섹션4")


# =============================================
# 검증 리포트
# =============================================
def verify_report(ws4, df26, df25):
    """검증 리포트 - 콘솔 + PDF 출력"""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    pdf_file = f"검증리포트_{MONTH}월_CLAUDE.pdf"

    # 한글 폰트 등록 (Windows 기본 폰트)
    font_name = "Helvetica"  # 기본값
    for path in ["C:/Windows/Fonts/malgun.ttf", "C:/Windows/Fonts/gulim.ttc", "C:/Windows/Fonts/batang.ttc"]:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("Korean", path))
                font_name = "Korean"
                break
            except: pass

    c = canvas.Canvas(pdf_file, pagesize=A4)
    width, height = A4
    y = height - 50

    def write(text, size=10, bold=False, indent=0):
        nonlocal y
        if y < 50:
            c.showPage()
            y = height - 50
        c.setFont(font_name, size)
        c.drawString(50 + indent, y, text)
        y -= size + 5

    # 헤더
    write(f"{MONTH}월 매출실적분석 검증 리포트", size=16)
    write("=" * 60, size=10)
    y -= 10

    print("\n" + "="*60)
    print(f"  {MONTH}월 검증 리포트")
    print("="*60)

    # 섹션1 합계 검증
    labels = {
        '수출$당월': (10,5), '수출$전년': (10,8), '수출$당해': (10,10),
        '내수당월':  (19,5), '내수전년':  (19,8), '내수당해':  (19,10),
    }
    write("[섹션1 소계 검증] 수식셀 값 vs 수동합산", size=12, bold=True)
    print("\n[섹션1 소계 검증] 수식셀 값 vs 수동합산")
    for label, (row, col) in labels.items():
        cell_val = ws4.cell(row, col).value or 0
        sum_val  = sum(ws4.cell(r, col).value or 0 for r in [row-3, row-2, row-1])
        diff = round(abs(cell_val - sum_val), 1)
        status = "OK" if diff < 1 else "X 차이있음"
        line = f"  {label:12s}: 셀={cell_val:>10.1f} | 합산={sum_val:>10.1f} | {status}"
        write(line, size=10, indent=10)
        print(f"  {label:12s}: 셀={cell_val:>10.1f} | 합산={sum_val:>10.1f} | {status}")

    y -= 10
    write("[섹션4 합섬 통합 검증] 통합 vs 한국+중국+비나", size=12, bold=True)
    print("\n[섹션4 합섬 통합 검증] 통합 vs 한국+중국+비나")
    for label, (tot_r, sub_rows, col) in [
        ('합섬매출당월', (116, [119,122,125], 4)),
        ('합섬이익당월', (117, [120,123,126], 4)),
        ('합섬매출당해', (116, [119,122,125], 7)),
        ('스텐매출당월', (128, [131,134],     4)),
    ]:
        tot_val = ws4.cell(tot_r, col).value or 0
        sub_val = sum(ws4.cell(r, col).value or 0 for r in sub_rows)
        diff = round(abs(tot_val - sub_val), 1)
        status = "OK" if diff < 1 else f"X 차이={diff}"
        line = f"  {label:14s}: 통합={tot_val:>10.1f} | 합산={sub_val:>10.1f} | {status}"
        write(line, size=10, indent=10)
        print(f"  {label:14s}: 통합={tot_val:>10.1f} | 합산={sub_val:>10.1f} | {status}")

    c.save()
    print(f"\n[완료] 검증 리포트 PDF 저장: {pdf_file}")


# =============================================
# 메인 실행
# =============================================
def run():
    print(f"{'='*60}")
    print(f"  {MONTH}월 매출실적분석 자동입력")
    print(f"{'='*60}")

    shutil.copy(TEMPLATE_FILE, OUTPUT_FILE)

    # 파일 로드
    wbp   = load_workbook(PLAN_FILE,     data_only=True)
    wbv   = load_workbook(VOL_FILE,      data_only=True)
    wbo   = load_workbook(OVERSEAS_FILE, data_only=True)
    wb_dsr   = load_workbook(DSR_FILE,   data_only=True)
    wb_steel = load_workbook(STEEL_FILE, data_only=True)
    wb4   = load_workbook(OUTPUT_FILE)

    wsp    = wbp.worksheets[0]
    ws_hs  = wbv['합섬종합']
    ws_st  = wbv['스텐종합']
    ws_jg  = wbv['제강종합']
    ws_cd  = wbo['청도26']
    ws_vn  = wbo['비나26']
    ws_dc  = wbo['DCV26']
    ws4    = wb4.worksheets[0]

    df26 = load_raw(RAW_26_FILE)
    df25 = load_raw(RAW_25_FILE)

    fill_section1(ws4, wsp, df26, df25, MONTH)
    fill_section2(ws4, ws_hs, ws_st, ws_jg, ws_cd, ws_vn, ws_dc, df26, df25, MONTH)
    fill_section3(ws4, ws_hs, ws_st, ws_jg, wbo, df26, df25, MONTH)
    fill_section4(ws4, wb_dsr, wb_steel, MONTH)

    print(f"\n[SKIP] 섹션1 기타매출(E/H/J) → 수동 입력")
    print(f"[SKIP] 환율(C15) → 수동 입력")

    wb4.save(OUTPUT_FILE)

    # 검증
    wb4_check = load_workbook(OUTPUT_FILE, data_only=True)
    verify_report(wb4_check.worksheets[0], df26, df25)

# ── 코멘트 자동 생성 ──────────────────────────────────────
    from comment_generator import generate_full_report, generate_full_report_with_customers, load_customer_data
    df26['월'] = df26['요청월'].astype(str).str[-2:] + '월'
    df26['부문'] = df26['부문'].replace('STS', '스텐')  # 추가
    months = sorted(df26['월'].unique())
    try:
        cdf = load_customer_data(CUSTOMER_FILES)
        report = generate_full_report_with_customers(df26, cdf, months)
        print("[완료] 거래처 상세 데이터 반영")
    except Exception as e:
        print(f"[경고] 거래처 상세 데이터 반영 실패: {e}")
        report = generate_full_report(df26, months)
    comment_file = f"코멘트_{MONTH}월.md"
    with open(comment_file, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"[완료] 코멘트 초안 저장: {comment_file}")

    print(f"\n저장 완료: {OUTPUT_FILE}")


if __name__ == "__main__":
    run()
