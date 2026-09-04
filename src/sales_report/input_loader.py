from __future__ import annotations

from collections import Counter, defaultdict
from math import log10
from pathlib import Path
from sys import intern

from openpyxl import load_workbook

from .models import CumulativeDataset, InputManifest, RawSheetCandidate, RecipientDataset
from .utils import normalize_period, number, text


CUMULATIVE_REQUIRED_HEADERS = {
    "부문",
    "내수/수출",
    "요청월",
    "레벨1명",
    "레벨2명",
    "레벨3명",
    "계정",
    "중량",
    "달러금액",
    "한국원화금액",
}
RECIPIENT_REQUIRED_HEADERS = {
    "출고요청년월",
    "LVL1NM",
    "LVL2NM",
    "LVL3NM",
    "계정",
    "원화금액",
    "달러금액",
    "거래처",
    "인수처명",
}

DIVISION_ALIASES = {
    "합섬": "synthetic",
    "STS": "stainless",
    "스텐": "stainless",
    "제강": "steel",
}
MARKET_ALIASES = {"내수": "domestic", "수출": "export"}
COMPANY_PREFIXES = {"합섬": "synthetic", "스텐": "stainless", "스틸": "steel"}


def _find_header(
    path: Path,
    required_headers: set[str],
    period_header: str | None = None,
    relevant_years: set[int] | None = None,
) -> RawSheetCandidate | None:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        candidates: list[tuple[tuple[int, int, int, int], RawSheetCandidate]] = []
        for sheet_position, worksheet in enumerate(workbook.worksheets):
            for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=15, values_only=True), 1):
                headers = [text(value) for value in row]
                if required_headers.issubset(set(headers)):
                    candidate = RawSheetCandidate(path, worksheet.title, row_index, headers)
                    relevant_period_rows = 0
                    valid_period_rows = 0
                    if period_header and period_header in headers:
                        period_column = headers.index(period_header) + 1
                        for period_row in worksheet.iter_rows(
                            min_row=row_index + 1,
                            min_col=period_column,
                            max_col=period_column,
                            values_only=True,
                        ):
                            period = normalize_period(period_row[0])
                            if len(period) != 6 or not period.isdigit():
                                continue
                            valid_period_rows += 1
                            if not relevant_years or int(period[:4]) in relevant_years:
                                relevant_period_rows += 1
                    raw_data_bonus = 1 if "RAWDATA" in "".join(worksheet.title.upper().split()) else 0
                    data_rows = max(0, (worksheet.max_row or 0) - row_index)
                    # 관련 연도의 실제 거래행이 많은 원자료를 우선하고, 동률이면 Raw data 시트를 선택한다.
                    score = (relevant_period_rows, valid_period_rows, raw_data_bonus, data_rows - sheet_position)
                    candidates.append((score, candidate))
                    break
        return max(candidates, key=lambda item: item[0])[1] if candidates else None
    finally:
        workbook.close()


def is_official_workbook(path: Path) -> bool:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(min_row=1, max_row=12, max_col=12, values_only=True):
                if any("사업계획대비" in text(value) and "매출실적" in text(value) for value in row):
                    return True
        return False
    finally:
        workbook.close()


def discover_official(input_dir: Path) -> Path:
    files = sorted(path for path in input_dir.glob("*.xlsx") if not path.name.startswith("~$"))
    if not files:
        raise ValueError("input 폴더에 .xlsx 파일이 없습니다.")
    official_files = [path for path in files if is_official_workbook(path)]
    if len(official_files) != 1:
        names = ", ".join(path.name for path in official_files) or "없음"
        raise ValueError(f"공식 실적표는 정확히 1개여야 합니다. 현재 감지: {names}")
    return official_files[0]


def _candidate_years(candidate: RawSheetCandidate, period_header: str) -> Counter[int]:
    workbook = load_workbook(candidate.path, read_only=True, data_only=True, keep_links=False)
    worksheet = workbook[candidate.sheet_name]
    index = {name: column for column, name in enumerate(candidate.headers) if name}
    years: Counter[int] = Counter()
    try:
        rows = worksheet.iter_rows(values_only=True)
        for _ in range(candidate.header_row):
            next(rows, None)
        period_column = index[period_header]
        for row in rows:
            period = normalize_period(row[period_column] if period_column < len(row) else None)
            if len(period) == 6 and period.isdigit():
                years[int(period[:4])] += 1
    finally:
        workbook.close()
    return years


def _classify_recipient(candidate: RawSheetCandidate) -> str:
    workbook = load_workbook(candidate.path, read_only=True, data_only=True, keep_links=False)
    worksheet = workbook[candidate.sheet_name]
    index = {name: column for column, name in enumerate(candidate.headers) if name}
    company_counts: Counter[str] = Counter()
    dollar_total = 0.0
    try:
        rows = worksheet.iter_rows(values_only=True)
        for _ in range(candidate.header_row):
            next(rows, None)
        for row_number, row in enumerate(rows, 1):
            lvl1 = text(row[index["LVL1NM"]] if index["LVL1NM"] < len(row) else None)
            for prefix, company in COMPANY_PREFIXES.items():
                if lvl1.startswith(prefix):
                    company_counts[company] += 1
            dollar_total += abs(number(row[index["달러금액"]] if index["달러금액"] < len(row) else None))
            if row_number >= 4000:
                break
    finally:
        workbook.close()
    if not company_counts:
        raise ValueError(f"인수처 파일의 부문을 판별할 수 없습니다: {candidate.path.name}")
    company = company_counts.most_common(1)[0][0]
    market = "export" if dollar_total > 0 else "domestic"
    return f"{company}_{market}"


def discover_inputs(input_dir: Path, official_path: Path, target_year: int, config: dict) -> InputManifest:
    cumulative: dict[int, RawSheetCandidate] = {}
    recipients: dict[tuple[str, int], RawSheetCandidate] = {}
    ignored: list[Path] = []
    relevant_years = {target_year - 1, target_year}

    files = sorted(path for path in input_dir.glob("*.xlsx") if not path.name.startswith("~$") and path != official_path)
    for path in files:
        cumulative_candidate = _find_header(path, CUMULATIVE_REQUIRED_HEADERS, "요청월", relevant_years)
        if cumulative_candidate:
            years = set(_candidate_years(cumulative_candidate, "요청월")) & relevant_years
            if not years:
                ignored.append(path)
                continue
            for year in years:
                if year in cumulative and cumulative[year].path != path:
                    raise ValueError(f"{year}년 누계파일이 2개 이상입니다: {cumulative[year].path.name}, {path.name}")
                cumulative[year] = cumulative_candidate
            continue

        recipient_candidate = _find_header(path, RECIPIENT_REQUIRED_HEADERS, "출고요청년월", relevant_years)
        if recipient_candidate:
            segment = _classify_recipient(recipient_candidate)
            if segment not in config["segments"]:
                ignored.append(path)
                continue
            years = set(_candidate_years(recipient_candidate, "출고요청년월")) & relevant_years
            if not years:
                ignored.append(path)
                continue
            for year in years:
                key = (segment, year)
                if key in recipients and recipients[key].path != path:
                    label = config["segments"][segment]["label"]
                    raise ValueError(f"{year}년 {label} 인수처 파일이 2개 이상입니다: {recipients[key].path.name}, {path.name}")
                recipients[key] = recipient_candidate
            continue

        ignored.append(path)

    missing_cumulative = relevant_years - set(cumulative)
    if missing_cumulative:
        raise ValueError(f"누계파일이 없습니다: {', '.join(str(year) for year in sorted(missing_cumulative))}년")

    missing_recipients = []
    for segment in config["report_order"]:
        for year in sorted(relevant_years):
            if (segment, year) not in recipients:
                missing_recipients.append(f"{year}년 {config['segments'][segment]['label']}")
    if missing_recipients:
        raise ValueError("다음 인수처 파일이 없습니다: " + ", ".join(missing_recipients))

    return InputManifest(official_path, cumulative, recipients, ignored)


def _cell(row, index: dict[str, int], name: str):
    column = index.get(name)
    return row[column] if column is not None and column < len(row) else None


def _intern(value: str) -> str:
    return intern(value)


def _display(value: str, config: dict) -> str:
    return config.get("display_names", {}).get(value, value) or "(미분류)"


def _segment_key(division: str, market: str) -> str | None:
    company = DIVISION_ALIASES.get(division)
    normalized_market = MARKET_ALIASES.get(market)
    return f"{company}_{normalized_market}" if company and normalized_market else None


def load_cumulative_datasets(manifest: InputManifest, target_period: str, config: dict) -> dict[str, CumulativeDataset]:
    target_year, target_month = int(target_period[:4]), int(target_period[4:])
    rows_by_segment: defaultdict[str, list[dict]] = defaultdict(list)
    source_rows: defaultdict[str, defaultdict[tuple[int, str, str], int]] = defaultdict(lambda: defaultdict(int))

    for source_year, candidate in sorted(manifest.cumulative_files.items()):
        workbook = load_workbook(candidate.path, read_only=True, data_only=True, keep_links=False)
        worksheet = workbook[candidate.sheet_name]
        index = {name: column for column, name in enumerate(candidate.headers) if name}
        rows = worksheet.iter_rows(values_only=True)
        for _ in range(candidate.header_row):
            next(rows, None)
        try:
            for raw in rows:
                period = normalize_period(_cell(raw, index, "요청월"))
                if len(period) != 6 or not period.isdigit() or int(period[:4]) != source_year:
                    continue
                year, month = int(period[:4]), int(period[4:])
                if year not in {target_year - 1, target_year} or month > target_month:
                    continue
                segment = _segment_key(text(_cell(raw, index, "부문")), text(_cell(raw, index, "내수/수출")))
                if not segment or segment not in config["segments"]:
                    continue
                account = text(_cell(raw, index, "계정"))
                if account in {str(value) for value in config.get("excluded_accounts", [6])}:
                    continue
                segment_config = config["segments"][segment]
                level1 = text(_cell(raw, index, "레벨1명"))
                level2 = text(_cell(raw, index, "레벨2명"))
                level3 = text(_cell(raw, index, "레벨3명"))
                levels = {1: level1, 2: level2, 3: level3}
                raw_amount = number(_cell(raw, index, "달러금액")) if segment_config["market"] == "export" else number(_cell(raw, index, "한국원화금액"))
                amount = raw_amount / (1000.0 if segment_config["market"] == "export" else 1_000_000.0)
                merchandise = any(word in level1 for word in config.get("product_weight_exclusions", {}).get("level1_contains", ["상품"]))
                weight = 0.0 if merchandise else number(_cell(raw, index, "중량")) / 1000.0
                primary = _intern(_display(levels.get(int(segment_config["primary_level"]), ""), config))
                secondary = _intern(_display(levels.get(int(segment_config["secondary_level"]), ""), config))
                region = text(_cell(raw, index, "담당자명")).strip() or "(미분류)"
                subregion = text(_cell(raw, index, "담당자(세부)명")).strip() or region
                rows_by_segment[segment].append(
                    {
                        "period": _intern(period),
                        "year": year,
                        "month": month,
                        "amount": amount,
                        "weight": weight,
                        "primary": primary,
                        "secondary": secondary,
                        "region": _intern(region),
                        "subregion": _intern(subregion),
                    }
                )
                source_rows[segment][(source_year, candidate.path.name, candidate.sheet_name)] += 1
        finally:
            workbook.close()

    datasets = {}
    for segment in config["report_order"]:
        sources = [
            {"role": "누계 수치", "year": year, "file": filename, "sheet": sheet, "rows": count}
            for (year, filename, sheet), count in sorted(source_rows[segment].items())
        ]
        datasets[segment] = CumulativeDataset(segment, config["segments"][segment]["label"], rows_by_segment[segment], sources)
    return datasets


def _recipient_name_key(value: str) -> str:
    return "".join(value.upper().split())


RECIPIENT_NAME_ALIASES = {
    _recipient_name_key("현대삼호중공업(주)"): "에이치디현대삼호㈜",
    _recipient_name_key("현대삼호중공업㈜"): "에이치디현대삼호㈜",
    _recipient_name_key("에이치디현대삼호(주)"): "에이치디현대삼호㈜",
    _recipient_name_key("에이치디현대삼호㈜"): "에이치디현대삼호㈜",
}


def _normalize_recipient_name(value: str) -> str:
    normalized = value.strip() or "(미지정)"
    return RECIPIENT_NAME_ALIASES.get(_recipient_name_key(normalized), normalized)


def _reference_value(reference_metrics, segment: str, year: int, target_year: int, value_type: str) -> float | None:
    if not reference_metrics or segment not in reference_metrics:
        return None
    metric = reference_metrics[segment]
    prefix = "prior_ytd" if year == target_year - 1 else "current_ytd"
    value = number(getattr(metric, f"{prefix}_{value_type}", None))
    return value if abs(value) > 1e-9 else None


def _closest_divisor(raw_total: float, reference_total: float | None, candidates: tuple[float, ...]) -> float | None:
    if reference_total is None or abs(raw_total) <= 1e-9:
        return None
    reference = abs(reference_total)
    raw = abs(raw_total)
    return min(candidates, key=lambda divisor: abs(log10(max(raw / divisor, 1e-12) / reference)))


def _fallback_amount_divisor(market: str, raw_values: list[float]) -> float:
    positives = sorted(abs(value) for value in raw_values if abs(value) > 1e-9)
    if not positives:
        return 1.0
    percentile_index = min(len(positives) - 1, int((len(positives) - 1) * 0.9))
    percentile_90 = positives[percentile_index]
    if market == "domestic":
        return 1_000_000.0 if percentile_90 >= 100_000 else 1.0
    return 1_000.0 if percentile_90 >= 1_000 else 1.0


def _unit_divisors(
    segment: str,
    market: str,
    year: int,
    target_year: int,
    raw_amounts: list[float],
    raw_weights: list[float],
    reference_metrics=None,
) -> tuple[float, float]:
    raw_amount_total = sum(raw_amounts)
    raw_weight_total = sum(raw_weights)
    amount_reference = _reference_value(reference_metrics, segment, year, target_year, "amount")
    weight_reference = _reference_value(reference_metrics, segment, year, target_year, "weight")

    amount_candidates = (1.0, 1_000_000.0) if market == "domestic" else (1.0, 1_000.0)
    amount_divisor = _closest_divisor(raw_amount_total, amount_reference, amount_candidates)
    if amount_divisor is None:
        amount_divisor = _fallback_amount_divisor(market, raw_amounts)

    weight_divisor = _closest_divisor(raw_weight_total, weight_reference, (1.0, 1_000.0))
    if weight_divisor is None:
        # 거래 원자료는 금액과 중량이 함께 원단위 또는 보고단위로 저장되는 구조를 기본으로 한다.
        weight_divisor = 1_000.0 if amount_divisor > 1.0 else 1.0
    return amount_divisor, weight_divisor


def load_recipient_datasets(
    manifest: InputManifest,
    target_period: str,
    config: dict,
    reference_metrics=None,
) -> dict[str, RecipientDataset]:
    target_year, target_month = int(target_period[:4]), int(target_period[4:])
    datasets = {}
    excluded_accounts = {str(value) for value in config.get("excluded_accounts", [6])}

    for segment in config["report_order"]:
        segment_config = config["segments"][segment]
        recipient_exclusions = segment_config.get("recipient_analysis_exclusions", {})
        excluded_level2 = {str(value).strip().upper() for value in recipient_exclusions.get("level2", [])}
        excluded_item_code_2 = {str(value).strip().upper() for value in recipient_exclusions.get("item_code_2", [])}
        collected: list[dict] = []
        sources: list[dict] = []
        for year in (target_year - 1, target_year):
            candidate = manifest.recipient_files[(segment, year)]
            workbook = load_workbook(candidate.path, read_only=True, data_only=True, keep_links=False)
            worksheet = workbook[candidate.sheet_name]
            index = {name: column for column, name in enumerate(candidate.headers) if name}
            rows = worksheet.iter_rows(values_only=True)
            for _ in range(candidate.header_row):
                next(rows, None)
            used_rows = 0
            pending_rows: list[dict] = []
            raw_amounts: list[float] = []
            raw_weights: list[float] = []
            try:
                for raw in rows:
                    period = normalize_period(_cell(raw, index, "출고요청년월"))
                    if len(period) != 6 or not period.isdigit() or int(period[:4]) != year or int(period[4:]) > target_month:
                        continue
                    if text(_cell(raw, index, "계정")) in excluded_accounts:
                        continue
                    level1 = text(_cell(raw, index, "LVL1NM"))
                    level2 = text(_cell(raw, index, "LVL2NM"))
                    level3 = text(_cell(raw, index, "LVL3NM"))
                    item_code_2 = text(_cell(raw, index, "품번2"))
                    if level2.strip().upper() in excluded_level2 or item_code_2.strip().upper() in excluded_item_code_2:
                        continue
                    levels = {1: level1, 2: level2, 3: level3}
                    raw_amount = number(_cell(raw, index, "달러금액")) if segment_config["market"] == "export" else number(_cell(raw, index, "원화금액"))
                    weight_header = "중량(Kg)" if "중량(Kg)" in index else "중량"
                    merchandise = any(
                        word in level1
                        for word in config.get("product_weight_exclusions", {}).get("level1_contains", ["상품"])
                    )
                    raw_weight = 0.0 if merchandise else number(_cell(raw, index, weight_header))
                    recipient_name = _normalize_recipient_name(
                        text(_cell(raw, index, "인수처명")) or text(_cell(raw, index, "거래처")) or "(미지정)"
                    )
                    primary = _intern(_display(levels.get(int(segment_config["primary_level"]), ""), config))
                    secondary = _intern(_display(levels.get(int(segment_config["secondary_level"]), ""), config))
                    pending_rows.append(
                        {
                            "period": _intern(period),
                            "year": year,
                            "month": int(period[4:]),
                            "raw_amount": raw_amount,
                            "raw_weight": raw_weight,
                            "primary": primary,
                            "secondary": secondary,
                            "recipient_key": _intern(_recipient_name_key(recipient_name)),
                            "recipient_name": _intern(recipient_name),
                        }
                    )
                    raw_amounts.append(raw_amount)
                    raw_weights.append(raw_weight)
                    used_rows += 1
            finally:
                workbook.close()
            amount_divisor, weight_divisor = _unit_divisors(
                segment,
                segment_config["market"],
                year,
                target_year,
                raw_amounts,
                raw_weights,
                reference_metrics,
            )
            for row in pending_rows:
                row["amount"] = row.pop("raw_amount") / amount_divisor
                row["weight"] = row.pop("raw_weight") / weight_divisor
            collected.extend(pending_rows)
            sources.append(
                {
                    "role": "인수처 분석",
                    "year": year,
                    "file": candidate.path.name,
                    "sheet": candidate.sheet_name,
                    "rows": used_rows,
                    "amount_divisor": amount_divisor,
                    "weight_divisor": weight_divisor,
                }
            )
        datasets[segment] = RecipientDataset(segment, segment_config["label"], collected, sources)
    return datasets
