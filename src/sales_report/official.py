from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .models import OfficialMetric, OfficialReport
from .utils import compact, number, text


DIVISION_TO_COMPANY = {"합섬": "synthetic", "스텐": "stainless", "제강": "steel"}


def _segment_key(division: str, market: str) -> str:
    return f"{DIVISION_TO_COMPANY[division]}_{market}"


def _find_title_and_sheet(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=12, max_col=12, values_only=True):
            for value in row:
                value_text = text(value)
                if "사업계획대비" in value_text and "매출실적" in value_text:
                    match = re.search(r"(20\d{2})년.*?(\d{1,2})월", value_text)
                    if not match:
                        wb.close()
                        raise ValueError(f"공식 실적표 제목에서 연월을 찾지 못했습니다: {value_text}")
                    return wb, ws, int(match.group(1)), int(match.group(2))
    wb.close()
    raise ValueError(f"공식 실적표 제목을 찾지 못했습니다: {path.name}")


def parse_official_report(path: Path, config: dict) -> OfficialReport:
    wb, ws, year, month = _find_title_and_sheet(path)
    sheet_name = ws.title
    metrics: dict[str, OfficialMetric] = {}

    # 판매금액: A열의 병합된 구분명을 아래 행까지 이어서 읽는다.
    amount_context = ""
    in_amount_section = False
    for row_idx in range(1, (ws.max_row or 200) + 1):
        a = compact(ws.cell(row_idx, 1).value)
        b = text(ws.cell(row_idx, 2).value)
        if "1.전체매출실적" in a:
            in_amount_section = True
            amount_context = ""
            continue
        if "2.제품판매량" in a:
            break
        if not in_amount_section:
            continue
        if a:
            amount_context = a
        market = None
        if amount_context in {"수출판매($)", "수출판매(＄)"}:
            market = "export"
        elif amount_context == "내수판매":
            market = "domestic"
        if market and b in DIVISION_TO_COMPANY:
            segment = _segment_key(b, market)
            segment_cfg = config["segments"].get(segment)
            if not segment_cfg:
                continue
            metrics[segment] = OfficialMetric(
                segment=segment,
                label=segment_cfg["label"],
                market=market,
                currency_unit="천달러" if market == "export" else "백만원",
                previous_amount=number(ws.cell(row_idx, 3).value),
                current_plan_amount=number(ws.cell(row_idx, 4).value),
                current_amount=number(ws.cell(row_idx, 5).value),
                prior_ytd_amount=number(ws.cell(row_idx, 8).value),
                current_ytd_plan_amount=number(ws.cell(row_idx, 9).value),
                current_ytd_amount=number(ws.cell(row_idx, 10).value),
            )

    # 제품판매중량: 상품을 제외한 공식 중량을 사용한다.
    weight_context = ""
    in_weight_section = False
    for row_idx in range(1, (ws.max_row or 200) + 1):
        a = compact(ws.cell(row_idx, 1).value)
        b = text(ws.cell(row_idx, 2).value)
        if "2.제품판매량" in a:
            in_weight_section = True
            weight_context = ""
            continue
        if "3.제품판매단가" in a:
            break
        if not in_weight_section:
            continue
        if a:
            weight_context = a
        market = weight_context if weight_context in {"수출", "내수"} else None
        if market and b in DIVISION_TO_COMPANY:
            segment = _segment_key(b, "export" if market == "수출" else "domestic")
            metric = metrics.get(segment)
            if not metric:
                continue
            metric.previous_weight = number(ws.cell(row_idx, 3).value)
            metric.current_plan_weight = number(ws.cell(row_idx, 4).value)
            metric.current_weight = number(ws.cell(row_idx, 5).value)
            metric.prior_ytd_weight = number(ws.cell(row_idx, 8).value)
            metric.current_ytd_plan_weight = number(ws.cell(row_idx, 9).value)
            metric.current_ytd_weight = number(ws.cell(row_idx, 10).value)

    wb.close()
    required_segments = set(config["report_order"])
    missing = required_segments - set(metrics)
    if missing:
        labels = [config["segments"][key]["label"] for key in sorted(missing)]
        raise ValueError(f"공식 실적표에서 다음 부문을 찾지 못했습니다: {', '.join(labels)}")
    return OfficialReport(path=path, year=year, month=month, metrics=metrics, sheet_name=sheet_name)
