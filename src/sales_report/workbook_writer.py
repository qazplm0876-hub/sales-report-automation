from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "17365D"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF3F8"
GREEN = "E2F0D9"
RED = "FCE4D6"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"
THIN_GRAY = Side(style="thin", color="D9E1F2")


def _base_sheet(ws, title: str, subtitle: str, columns: int):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(columns, 2))
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(name="맑은 고딕", size=16, bold=True, color=WHITE)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(columns, 2))
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).font = Font(name="맑은 고딕", size=9, color="666666")
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 30


def _write_table(ws, headers: list[str], rows: list[list], start_row: int = 4):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="맑은 고딕", size=9, bold=True, color="1F1F1F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    for row_idx, values in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = Border(bottom=THIN_GRAY)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.alignment = Alignment(horizontal="right", vertical="center")
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    ws.freeze_panes = f"A{start_row + 1}"
    ws.row_dimensions[start_row].height = 34


def _set_widths(ws, widths: list[float]):
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = min(width, 80)


def _format_numeric_columns(ws, columns: list[int], start_row: int, end_row: int, fmt: str):
    for col in columns:
        for row in range(start_row, end_row + 1):
            ws.cell(row, col).number_format = fmt


def _color_delta_columns(ws, columns: list[int], start_row: int, end_row: int):
    if end_row < start_row:
        return
    for col in columns:
        letter = get_column_letter(col)
        target = f"{letter}{start_row}:{letter}{end_row}"
        ws.conditional_formatting.add(target, CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=GREEN)))
        ws.conditional_formatting.add(target, CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=RED)))


def _add_summary_sheet(wb: Workbook, analysis: dict, official):
    ws = wb.active
    ws.title = "요약"
    headers = ["비교", "부문", "단위", "기준 판매금액", "당기 판매금액", "증감액", "증감률", "기준 제품판매중량", "당기 제품판매중량", "중량 증감", "중량 증감률", "기준 평균단가", "당기 평균단가", "단가 증감률", "중량 효과", "단가·구성 효과"]
    rows = []
    for row in analysis["summary"]:
        rows.append([row["comparison"], row["label"], row["currency_unit"], row["previous_amount"], row["current_amount"], row["amount_delta"], row["amount_pct"], row["previous_weight"], row["current_weight"], row["weight_delta"], row["weight_pct"], row["previous_price"], row["current_price"], row["price_pct"], row["volume_effect"], row["price_mix_effect"]])
    subtitle = f"MODEL STATUS: {analysis['model_status']} | 기준월: {analysis['target_period']} | 생성: {datetime.now():%Y-%m-%d %H:%M} | KPI는 누계파일 계산 후 공식표 검산"
    _base_sheet(ws, f"{analysis['year']}년 {analysis['month']}월 매출실적 분석", subtitle, len(headers))
    _write_table(ws, headers, rows)
    _set_widths(ws, [12, 15, 10, 16, 16, 14, 11, 18, 18, 14, 12, 14, 14, 12, 14, 16])
    end = 4 + len(rows)
    _format_numeric_columns(ws, [4, 5, 6, 8, 9, 10, 15, 16], 5, end, '#,##0.0;[Red](#,##0.0);-')
    _format_numeric_columns(ws, [7, 11, 14], 5, end, '0.0%;[Red](0.0%);-')
    _format_numeric_columns(ws, [12, 13], 5, end, '0.00')
    _color_delta_columns(ws, [6, 7, 10, 11, 14, 15, 16], 5, end)


def _simple_sheet(wb: Workbook, name: str, title: str, subtitle: str, headers: list[str], rows: list[list], widths: list[float], number_columns=None, pct_columns=None, delta_columns=None):
    ws = wb.create_sheet(name)
    _base_sheet(ws, title, subtitle, len(headers))
    _write_table(ws, headers, rows)
    _set_widths(ws, widths)
    end = 4 + len(rows)
    if number_columns:
        _format_numeric_columns(ws, number_columns, 5, end, '#,##0.0;[Red](#,##0.0);-')
    if pct_columns:
        _format_numeric_columns(ws, pct_columns, 5, end, '0.0%;[Red](0.0%);-')
    if delta_columns:
        _color_delta_columns(ws, delta_columns, 5, end)
    return ws


def write_analysis_workbook(path: Path, analysis: dict, official, config: dict):
    wb = Workbook()
    wb.properties.title = f"{analysis['year']}년 {analysis['month']}월 매출실적 분석"
    wb.properties.subject = "사업계획대비 매출실적 자동 분석"
    wb.properties.creator = "기획팀 박상원"
    _add_summary_sheet(wb, analysis, official)

    role_values = [[row["source"], row["used_for"], row["not_used_for"]] for row in analysis["data_roles"]]
    ws_roles = _simple_sheet(
        wb,
        "자료역할",
        "입력자료 역할",
        "자료별 용도를 분리해 인수처 파일의 불완전한 합계가 전체 KPI에 섞이지 않도록 했습니다.",
        ["자료", "사용 용도", "사용하지 않는 용도"],
        role_values,
        [30, 75, 48],
    )
    for row in range(5, 5 + len(role_values)):
        ws_roles.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws_roles.cell(row, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws_roles.row_dimensions[row].height = 42

    checks = sorted(analysis["checks"], key=lambda row: (row["status"] != "WARN", config["report_order"].index(row["segment"]), row["check"]))
    ws_checks = _simple_sheet(
        wb, "검산", "자료 검산", "WARN 항목을 먼저 확인한 뒤 보고문안을 사용하세요.",
        ["상태", "부문", "검산항목", "공식표", "누계파일 계산", "차이", "일치율", "단위", "확인 위치", "비고"],
        [[row["status"], row["label"], row["check"], row["official"], row["calculated"], row["difference"], row["match_pct"], row["unit"], row["where_to_fix"], row["notes"]] for row in checks],
        [10, 15, 22, 14, 14, 14, 12, 10, 35, 40], [4, 5, 6], [7], [6, 7],
    )
    for row in range(5, 5 + len(checks)):
        status = ws_checks.cell(row, 1).value
        ws_checks.cell(row, 1).fill = PatternFill("solid", fgColor=RED if status == "WARN" else GREEN)
        ws_checks.cell(row, 1).font = Font(name="맑은 고딕", size=9, bold=True)

    item_headers = ["비교", "부문", "순위", "품목", "기준금액", "당기금액", "증감액", "증감률", "기준중량", "당기중량", "중량증감", "중량증감률", "기준단가", "당기단가", "단가증감률"]
    item_values = [[row["comparison"], row["label"], row["rank"], row["primary"], row["previous_amount"], row["current_amount"], row["amount_delta"], row["amount_pct"], row["previous_weight"], row["current_weight"], row["weight_delta"], row["weight_pct"], row["previous_price"], row["current_price"], row["price_pct"]] for row in analysis["items"]]
    ws_items = _simple_sheet(wb, "품목증감", "품목별 증감", "합섬·스텐 내수·수출은 레벨1명, 제강은 레벨2명 기준입니다. 누계파일에서 계산하며 SR1·SR2 명칭을 적용했습니다.", item_headers, item_values, [12, 15, 8, 24, 14, 14, 14, 11, 14, 14, 14, 12, 13, 13, 12], [5, 6, 7, 9, 10, 11, 13, 14], [8, 12, 15], [7, 8, 11, 12, 15])
    _format_numeric_columns(ws_items, [13, 14], 5, 4 + len(item_values), '0.00')

    secondary_values = [[row["comparison"], row["label"], row["primary"], row["secondary"], row["previous_amount"], row["current_amount"], row["amount_delta"], row["amount_pct"], row["previous_weight"], row["current_weight"], row["weight_delta"], row["weight_pct"], row["previous_price"], row["current_price"], row["price_pct"]] for row in analysis["secondary_items"]]
    ws_secondary = _simple_sheet(wb, "세부품목", "세부품목별 증감", "합섬·스텐 내수·수출의 레벨2명 기준이며 누계파일에서 계산합니다. 제강은 레벨2명에서 바로 인수처 분석으로 연결하므로 이 시트에서 제외합니다.", ["비교", "부문", "품목", "세부품목", "기준금액", "당기금액", "증감액", "증감률", "기준중량", "당기중량", "중량증감", "중량증감률", "기준단가", "당기단가", "단가증감률"], secondary_values, [12, 15, 22, 26, 14, 14, 14, 11, 14, 14, 14, 12, 13, 13, 12], [5, 6, 7, 9, 10, 11, 13, 14], [8, 12, 15], [7, 8, 11, 12, 15])
    _format_numeric_columns(ws_secondary, [13, 14], 5, 4 + len(secondary_values), '0.00')

    recipient_values = [[row["comparison"], row["label"], row["primary"], "" if row["primary"] == row["secondary"] else row["secondary"], row["rank"], row["recipient"], row["status"], row["previous_amount"], row["current_amount"], row["amount_delta"], row["amount_pct"], row["prior_same_month"], row["prior_ytd"], row["current_ytd"], row["prior_year_exists"]] for row in analysis["recipients"]]
    ws_rec = _simple_sheet(
        wb,
        "인수처증감",
        "인수처별 주요 증감",
        "이 시트의 금액은 인수처 파일 기준이며 전체 KPI·중량·평균단가 계산에는 사용하지 않습니다. 인수처명이 없을 때만 거래처명을 사용합니다.",
        ["비교", "부문", "품목", "세부품목", "순위", "인수처", "패턴", "기준금액", "당기금액", "증감액", "증감률", "전년동월금액", "전년동기누계", "당해누계", "전년실적"],
        recipient_values,
        [12, 15, 20, 24, 8, 34, 20, 14, 14, 14, 11, 15, 16, 16, 12],
        [8, 9, 10, 12, 13, 14],
        [11],
        [10, 11],
    )

    unit_by_segment = {
        row["segment"]: row["currency_unit"]
        for row in analysis["summary"]
        if row["comparison"] == "전월 대비"
    }
    mix_values = []
    for row in analysis["recipient_mix_shifts"]:
        unit = unit_by_segment[row["segment"]]
        increase_text = ", ".join(
            f"{item['recipient']} {item['amount_delta']:+,.0f}{unit}" for item in row["increase_drivers"]
        )
        decrease_text = ", ".join(
            f"{item['recipient']} {item['amount_delta']:+,.0f}{unit}" for item in row["decrease_drivers"]
        )
        divergence_text = "; ".join(
            f"{item['recipient']}: 금액 {item['amount_delta']:+,.0f}{unit}, 중량 {item['weight_delta']:+,.1f}톤, "
            f"단가 {item['previous_price']:.2f}→{item['current_price']:.2f}"
            for item in row["divergent_recipients"]
        )
        mix_values.append(
            [
                row["comparison"],
                row["label"],
                row["primary"],
                "" if row["primary"] == row["secondary"] else row["secondary"],
                row["product_amount_delta"],
                row["gross_increase"],
                -row["gross_decrease"],
                row["gross_movement"],
                row["net_share"],
                increase_text,
                decrease_text,
                divergence_text,
            ]
        )
    ws_mix = _simple_sheet(
        wb,
        "인수처구성변화",
        "인수처 구성 변화 감지",
        "품목 순증감은 작아도 인수처별 증가·감소가 동시에 큰 경우를 표시합니다. 금액·중량 역행은 평균단가 또는 제품 구성 변화 확인이 필요한 신호입니다.",
        ["비교", "부문", "품목", "세부품목", "품목 순증감", "인수처 증가합계", "인수처 감소합계", "총 변동", "순증감/총변동", "주요 증가 인수처", "주요 감소 인수처", "금액·중량 역행"],
        mix_values,
        [12, 15, 20, 24, 14, 16, 16, 14, 15, 52, 42, 70],
        [5, 6, 7, 8],
        [9],
        [5, 6, 7, 9],
    )
    for row_index in range(5, 5 + len(mix_values)):
        for column in (10, 11, 12):
            ws_mix.cell(row_index, column).alignment = Alignment(wrap_text=True, vertical="top")
        ws_mix.row_dimensions[row_index].height = 42

    region_values = [[row["comparison"], row["label"], row["rank"], row["region"], row["previous_amount"], row["current_amount"], row["amount_delta"], row["amount_pct"], row["previous_weight"], row["current_weight"], row["weight_delta"], row["weight_pct"], row["previous_price"], row["current_price"], row["price_pct"]] for row in analysis["regions"]]
    ws_region = _simple_sheet(wb, "지역증감", "수출 지역별 증감", "누계파일의 담당자명 지역 구분을 기준으로 계산합니다.", ["비교", "부문", "순위", "지역", "기준금액", "당기금액", "증감액", "증감률", "기준중량", "당기중량", "중량증감", "중량증감률", "기준단가", "당기단가", "단가증감률"], region_values, [12, 15, 8, 20, 14, 14, 14, 11, 14, 14, 14, 12, 13, 13, 12], [5, 6, 7, 9, 10, 11, 13, 14], [8, 12, 15], [7, 8, 11, 12, 15])
    _format_numeric_columns(ws_region, [13, 14], 5, 4 + len(region_values), '0.00')

    prior_month_columns = [f"{analysis['year'] - 1}{month:02d}" for month in range(1, analysis["month"] + 1)]
    current_month_columns = [f"{analysis['year']}{month:02d}" for month in range(1, analysis["month"] + 1)]
    month_columns = prior_month_columns + current_month_columns
    trend_headers = ["부문", "품목", "세부품목", "인수처", "전년동기누계", "당해누계", "누계증감"] + month_columns
    trend_values = [[row["label"], row["primary"], "" if row["primary"] == row["secondary"] else row["secondary"], row["recipient"], row["prior_ytd"], row["current_ytd"], row["ytd_delta"]] + [row.get(period, 0.0) for period in month_columns] for row in analysis["trends"]]
    ws_trend = _simple_sheet(wb, "인수처월별추이", "주요 인수처 월별 추이", "인수처 파일 기준 전년·당해 월별 판매금액입니다. 전년 실적 존재 여부와 신규·재개·미출고 판단에 사용합니다.", trend_headers, trend_values, [15, 22, 24, 34, 16, 16, 15] + [12] * len(month_columns), list(range(5, len(trend_headers) + 1)), None, [7])

    question_values = [[row["no"], row["segment"], row["type"], row["question"], row["basis"]] for row in analysis["questions"]]
    ws_q = _simple_sheet(wb, "추가확인사항", "담당자 추가 확인사항", "일회성·가격조정·출고재개 여부처럼 원자료만으로 단정할 수 없는 내용입니다.", ["번호", "부문", "유형", "확인사항", "근거"], question_values, [8, 15, 18, 80, 28])
    for row in range(5, 5 + len(question_values)):
        ws_q.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")
        ws_q.row_dimensions[row].height = 38

    source_values = [["공식 검산", analysis["year"], "전체", official.path.name, official.sheet_name, "-", analysis["target_period"]]] + [[row["role"], row["year"], row["label"], row["file"], row["sheet"], row["rows"], row["period"]] for row in analysis["sources"]]
    _simple_sheet(
        wb,
        "원본파일",
        "원본파일 및 자료 역할 기록",
        "누계 수치자료와 인수처 분석자료를 분리해 기록합니다. 금액 단위는 내수 백만원, 수출 천달러이며 중량은 톤입니다.",
        ["자료역할", "연도", "부문", "파일", "선택 시트", "분석 행수", "기준월"],
        source_values,
        [18, 10, 16, 48, 22, 14, 12],
        [2, 6],
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
